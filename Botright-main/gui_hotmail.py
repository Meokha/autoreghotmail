import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import random
import time
import asyncio
import queue
# pandas is optional for reading Excel files. Prefer pandas if available
try:
    import pandas as pd
except Exception:
    pd = None
    try:
        from openpyxl import load_workbook
    except Exception:
        load_workbook = None
from datetime import datetime

import botright
import hcaptcha_challenger as solver
from captcha_solver import warmup_account
from db import fetch_pending_accounts, fetch_all_emails, update_warmup_status
from hotmail_auto_simple import HotmailAccountCreator
from playwright._impl._errors import TargetClosedError


class MultiHotmailGUI:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Auto hotmail registration")
        self.window.geometry("640x460")
        self.running = False
        self.threads = []
        self.account_list: list[dict] = []
        # Lấy kích thước màn hình để tính layout cửa sổ trình duyệt
        # Gọi update_idletasks để Tk có thể trả về đúng thông số màn hình
        self.window.update_idletasks()
        self.screen_w = self.window.winfo_screenwidth()
        self.screen_h = self.window.winfo_screenheight()
        self.warmup_accounts: list[dict] = []
        self.warmup_running = False
        self.warmup_thread: threading.Thread | None = None
        self.warmup_email_map: dict[str, dict] = {}
        self.warmup_queue: queue.Queue | None = None
        self.warmup_total = 0
        self.warmup_completed = 0
        self.warmup_active_workers = 0
        self.warmup_lock = threading.Lock()
        self.warmup_email_ring: list[str] = []
        # Custom emails / proxies support
        from collections import deque

        self.custom_emails = deque()  # holds strings like 'local' or 'local@domain' or 'local@domain|proxy'
        self.custom_emails_lock = threading.Lock()
        self.use_custom_emails = tk.BooleanVar(value=False)

        self.create_proxies: list[str] = []
        self.create_proxy_index = 0
        self.create_proxy_lock = threading.Lock()

        self.warmup_proxies: list[str] = []
        self.warmup_proxy_index = 0
        self.warmup_proxy_lock = threading.Lock()
        self.create_widgets()

    def create_widgets(self):
        notebook = ttk.Notebook(self.window)
        notebook.pack(fill="both", expand=True)

        self.create_tab = ttk.Frame(notebook)
        self.warmup_tab = ttk.Frame(notebook)
        notebook.add(self.create_tab, text="Tạo tài khoản")
        notebook.add(self.warmup_tab, text="Nuôi tài khoản")

        self._build_create_tab()
        self._build_warmup_tab()

    # ===== TAB TẠO TÀI KHOẢN =====
    def _build_create_tab(self):
        top = ttk.Frame(self.create_tab)
        top.pack(fill="x", pady=10, padx=10)

        label = ttk.Label(top, text="Số tài khoản cần tạo:")
        label.pack(side="left")

        self.num_accounts = ttk.Entry(top, width=8)
        self.num_accounts.pack(side="left", padx=8)
        self.num_accounts.insert(0, "1")

        self.fast_mode = tk.BooleanVar()
        fast_check = ttk.Checkbutton(top, text="Fast Mode (Tối ưu tốc độ)", variable=self.fast_mode)
        fast_check.pack(side="left", padx=8)

        domain_frame = ttk.Frame(self.create_tab)
        domain_frame.pack(fill="x", padx=10, pady=5)
        ttk.Label(domain_frame, text="Domain:").pack(side="left")
        self.domain_choice = ttk.Combobox(domain_frame, values=["random", "hotmail", "outlook"], state="readonly", width=10)
        self.domain_choice.set("random")
        self.domain_choice.pack(side="left", padx=8)

        # Custom emails file picker
        custom_frame = ttk.Frame(self.create_tab)
        custom_frame.pack(fill="x", padx=10, pady=6)
        self.use_custom_cb = ttk.Checkbutton(custom_frame, text="Sử dụng email tùy chỉnh (file)", variable=self.use_custom_emails)
        self.use_custom_cb.pack(side="left")
        ttk.Button(custom_frame, text="Chọn file email...", command=self.load_custom_emails).pack(side="left", padx=6)
        self.custom_emails_label = ttk.Label(custom_frame, text="(0 email)" )
        self.custom_emails_label.pack(side="left", padx=6)

        # Proxies for creation
        proxy_create_frame = ttk.Frame(self.create_tab)
        proxy_create_frame.pack(fill="x", padx=10, pady=(0,6))
        ttk.Label(proxy_create_frame, text="Proxy cho tạo tài khoản:").pack(side="left")
        ttk.Button(proxy_create_frame, text="Chọn file proxy...", command=self.load_create_proxies).pack(side="left", padx=6)
        self.create_proxies_label = ttk.Label(proxy_create_frame, text="(0 proxy)")
        self.create_proxies_label.pack(side="left", padx=6)

        concurrency_frame = ttk.Frame(self.create_tab)
        concurrency_frame.pack(fill="x", padx=10, pady=5)
        ttk.Label(concurrency_frame, text="Số cửa sổ song song (1-3):").pack(side="left")
        self.concurrency_var = tk.IntVar(value=1)
        concurrency_spin = ttk.Spinbox(concurrency_frame, from_=1, to=3, width=5, textvariable=self.concurrency_var)
        concurrency_spin.pack(side="left", padx=8)

        # Password options
        pwd_frame = ttk.Frame(self.create_tab)
        pwd_frame.pack(fill="x", padx=10, pady=5)
        ttk.Label(pwd_frame, text="Password:").pack(side="left")
        self.password_mode = tk.StringVar(value="random")
        rb_random = ttk.Radiobutton(pwd_frame, text="Random mỗi tài khoản", value="random", variable=self.password_mode, command=self._toggle_password_entry)
        rb_fixed = ttk.Radiobutton(pwd_frame, text="Cố định cho tất cả", value="fixed", variable=self.password_mode, command=self._toggle_password_entry)
        rb_random.pack(side="left", padx=6)
        rb_fixed.pack(side="left", padx=6)
        self.fixed_password = ttk.Entry(pwd_frame, width=24, show="*")
        self.fixed_password.pack(side="left", padx=8)
        self.fixed_password.configure(state="disabled")

        button_frame = ttk.Frame(self.create_tab)
        button_frame.pack(pady=8)

        start_button = ttk.Button(button_frame, text="Bắt đầu", command=self.start_creation)
        start_button.pack(side="left", padx=5)

        stop_button = ttk.Button(button_frame, text="Stop", command=self.stop_creation)
        stop_button.pack(side="left", padx=5)

        delete_button = ttk.Button(button_frame, text="Delete", command=self.delete_selected)
        delete_button.pack(side="left", padx=5)

        self.tree = ttk.Treeview(
            self.create_tab,
            columns=("stt", "email", "password", "status"),
            show="headings",
            height=12,
        )
        self.tree.heading("stt", text="STT")
        self.tree.heading("email", text="Email")
        self.tree.heading("password", text="Password")
        self.tree.heading("status", text="Trạng thái")
        self.tree.column("stt", width=60, anchor="center")
        self.tree.column("email", width=260, anchor="w")
        self.tree.column("password", width=220, anchor="w")
        self.tree.column("status", width=120, anchor="center")
        self.tree.pack(pady=10, fill="both", expand=True, padx=10)

        status_frame = ttk.Frame(self.create_tab)
        status_frame.pack(fill="x", padx=10, pady=5)
        self.status_var = tk.StringVar(value="Sẵn sàng")
        ttk.Label(status_frame, textvariable=self.status_var).pack(side="left")

    # ===== TAB NUÔI TÀI KHOẢN =====
    def _build_warmup_tab(self):
        header = ttk.Label(self.warmup_tab, text="Nuôi tài khoản đã tạo (status = created)", font=("Segoe UI", 11, "bold"))
        header.pack(anchor="w", padx=10, pady=(10, 0))

        mode_frame = ttk.Frame(self.warmup_tab)
        mode_frame.pack(fill="x", padx=10, pady=4)
        ttk.Label(mode_frame, text="Chọn cách nuôi:").pack(side="left")
        self.warmup_mode = tk.StringVar(value="auto")
        ttk.Radiobutton(
            mode_frame,
            text="1) Nhập số lượng",
            value="auto",
            variable=self.warmup_mode,
        ).pack(side="left", padx=6)
        ttk.Radiobutton(
            mode_frame,
            text="2) Chọn tài khoản trong bảng",
            value="manual",
            variable=self.warmup_mode,
        ).pack(side="left", padx=6)

        control = ttk.Frame(self.warmup_tab)
        control.pack(fill="x", padx=10, pady=4)

        ttk.Label(control, text="Số tài khoản tải từ DB:").pack(side="left")
        self.warmup_limit_entry = ttk.Entry(control, width=6)
        self.warmup_limit_entry.pack(side="left", padx=6)

        ttk.Label(control, text="Cửa sổ song song (1-3):").pack(side="left", padx=(15, 4))
        self.warmup_concurrency_var = tk.IntVar(value=1)
        warmup_conc_spin = ttk.Spinbox(control, from_=1, to=3, width=4, textvariable=self.warmup_concurrency_var)
        warmup_conc_spin.pack(side="left")

        # Warmup proxy file picker
        warmup_proxy_frame = ttk.Frame(self.warmup_tab)
        warmup_proxy_frame.pack(fill="x", padx=10, pady=(6, 4))
        ttk.Label(warmup_proxy_frame, text="Proxy cho nuôi (tùy chọn):").pack(side="left")
        ttk.Button(warmup_proxy_frame, text="Chọn file proxy...", command=self.load_warmup_proxies).pack(side="left", padx=6)
        self.warmup_proxies_label = ttk.Label(warmup_proxy_frame, text="(0 proxy)")
        self.warmup_proxies_label.pack(side="left", padx=6)
        # Option to ignore proxies during warmup (even if accounts have proxy saved)
        # Default to False so warmup won't use proxies unless user enables them
        self.use_warmup_proxies = tk.BooleanVar(value=False)
        ttk.Checkbutton(warmup_proxy_frame, text="Sử dụng proxy khi nuôi", variable=self.use_warmup_proxies).pack(side="left", padx=8)
        # Fast warmup mode: reduce internal waits/timeouts to speed up the process
        self.warmup_fast = tk.BooleanVar(value=False)
        ttk.Checkbutton(warmup_proxy_frame, text="Fast nuôi (nhanh)", variable=self.warmup_fast).pack(side="left", padx=8)

        ttk.Label(control, text="Lọc trạng thái:").pack(side="left", padx=(15, 4))
        self.warmup_status_filter = ttk.Combobox(control, values=["created", "warmup_failed", "warmed", "all"], width=13, state="readonly")
        self.warmup_status_filter.set("created")
        self.warmup_status_filter.pack(side="left")

        button_frame = ttk.Frame(self.warmup_tab)
        button_frame.pack(fill="x", padx=10, pady=6)
        load_btn = ttk.Button(button_frame, text="Tải danh sách", command=self.load_warmup_accounts)
        load_btn.pack(side="left", padx=4)
        start_btn = ttk.Button(button_frame, text="Bắt đầu nuôi", command=self.start_warmup)
        start_btn.pack(side="left", padx=4)
        stop_btn = ttk.Button(button_frame, text="Dừng", command=self.stop_warmup)
        stop_btn.pack(side="left", padx=4)

        self.warmup_tree = ttk.Treeview(
            self.warmup_tab,
            columns=("stt", "email", "status", "last", "otp"),
            show="headings",
            height=12,
            selectmode="extended",
        )
        self.warmup_tree.heading("stt", text="STT")
        self.warmup_tree.heading("email", text="Email")
        self.warmup_tree.heading("status", text="Trạng thái")
        self.warmup_tree.heading("last", text="Hoạt động cuối")
        self.warmup_tree.heading("otp", text="OTP")
        self.warmup_tree.column("stt", width=60, anchor="center")
        self.warmup_tree.column("email", width=260, anchor="w")
        self.warmup_tree.column("status", width=120, anchor="center")
        self.warmup_tree.column("last", width=140, anchor="center")
        self.warmup_tree.column("otp", width=160, anchor="center")
        self.warmup_tree.pack(fill="both", expand=True, padx=10, pady=10)

        status_frame = ttk.Frame(self.warmup_tab)
        status_frame.pack(fill="x", padx=10, pady=(0, 10))
        self.warmup_status_var = tk.StringVar(value="Chưa tải danh sách")
        ttk.Label(status_frame, textvariable=self.warmup_status_var).pack(side="left")

    def add_account_to_table(self, email: str, password: str, status_text: str):
        stt = len(self.account_list) + 1
        self.account_list.append({"email": email, "password": password, "status": status_text})
        self.tree.insert("", "end", values=(stt, email, password, status_text))
        self.status_var.set(f"Đã xử lý {stt} tài khoản")

    def stop_creation(self):
        self.running = False
        self.status_var.set("Đang dừng...")
        print("⛔ Đang dừng quá trình tạo tài khoản...")

    def delete_selected(self):
        selected_items = self.tree.selection()
        if not selected_items:
            print("⚠️ Vui lòng chọn tài khoản cần xóa!")
            return
        for item in selected_items:
            values = self.tree.item(item)['values']
            if values:
                email = values[1]
                self.account_list = [item for item in self.account_list if item["email"] != email]
                self.tree.delete(item)
        for idx, item in enumerate(self.tree.get_children(), 1):
            self.tree.set(item, "stt", idx)

    def _toggle_password_entry(self):
        mode = self.password_mode.get()
        if mode == "fixed":
            self.fixed_password.configure(state="normal")
        else:
            self.fixed_password.delete(0, tk.END)
            self.fixed_password.configure(state="disabled")

    def _resolve_domain(self):
        choice = (self.domain_choice.get() or "random").lower()
        if choice == "hotmail":
            return "hotmail"
        if choice == "outlook":
            return "outlook"
        return "random"  # đánh dấu để random từng tài khoản

    def _abort_creation(self, message: str):
        if not self.running:
            return
        self.running = False
        formatted = f"FAIL: {message}"
        print(f"❌ {formatted}")
        self.window.after(0, self.status_var.set, formatted)

    def _calc_window_conf(self, columns_count: int, index: int) -> dict:
        pad = 10
        columns = max(1, min(3, columns_count))
        layout_columns = columns if columns > 1 else 3
        win_w = int((self.screen_w - (layout_columns + 1) * pad) / layout_columns)
        target_height = int(self.screen_h * 0.55)
        win_h = min(target_height, self.screen_h - 2 * pad, 620)
        win_w = max(win_w, 400)
        win_h = max(win_h, 355)
        total_used_w = win_w * layout_columns + pad * (layout_columns + 1)
        if total_used_w > self.screen_w:
            win_w = int((self.screen_w - (layout_columns + 1) * pad) / layout_columns)
        win_h = min(win_h, self.screen_h - 2 * pad)
        col = index % layout_columns
        x = pad + col * (win_w + pad)
        y = pad
        return {
            "viewport": {"width": win_w, "height": win_h},
            "extra_args": [f"--window-position={x},{y}", f"--window-size={win_w},{win_h}"],
        }

    def start_creation(self):
        try:
            # start_creation: proceed normally (file-reading checks are handled when loading files)
            if self.running:
                print("⚠️ Đang trong quá trình tạo tài khoản!")
                return
            total = int(self.num_accounts.get())
            if total <= 0:
                print("Số tài khoản phải lớn hơn 0")
                return
            # If user selected custom emails, prefer using that list size for total
            try:
                if self.use_custom_emails.get():
                    with self.custom_emails_lock:
                        available = len(self.custom_emails)
                    if available > 0:
                        # If user left default 1 but file contains many entries, create them all
                        if total <= 1 and available > 1:
                            total = available
                            # update UI field to reflect chosen total
                            try:
                                self.num_accounts.delete(0, tk.END)
                                self.num_accounts.insert(0, str(total))
                            except Exception:
                                pass
                        else:
                            total = min(total, available)

            except Exception:
                pass

            # Giới hạn tối đa 3 cửa sổ chạy song song (mặc định 1 để an toàn)
            try:
                requested_concurrency = int(self.concurrency_var.get())
            except (TypeError, ValueError):
                requested_concurrency = 1
            concurrency = max(1, min(3, requested_concurrency, total))

            self.running = True
            self.status_var.set(f"Đang tạo {total} tài khoản (tối đa {concurrency} cửa sổ song song)...")
            print("🔥 HOTMAIL AUTO CREATOR - GUI VERSION")
            print("=" * 60)
            print(f"Bắt đầu tạo {total} tài khoản Hotmail...")
            print("=" * 60)

            fast_mode_val = self.fast_mode.get()
            chosen_domain = self._resolve_domain()
            if chosen_domain == "random":
                print("Domain: sẽ random @hotmail/@outlook cho từng tài khoản")
            else:
                print(f"Domain cố định: @{chosen_domain}.com")
            mode = self.password_mode.get()
            fixed_pwd_value = None
            if mode == "fixed":
                val = (self.fixed_password.get() or "").strip()
                if not val:
                    print("⚠️ Bạn chọn dùng mật khẩu cố định nhưng chưa nhập. Sẽ dùng random.")
                else:
                    fixed_pwd_value = val

            # Pre-install solver models once to avoid parallel file locks
            try:
                solver.install(upgrade=False)
            except Exception:
                pass

            # Bộ đếm công việc dùng chung giữa các worker
            lock = threading.Lock()
            # If we're using a custom email list and user wants to only create those,
            # prefer to process exactly that many and do NOT generate extra addresses.
            only_custom = False
            try:
                if self.use_custom_emails.get():
                    with self.custom_emails_lock:
                        available = len(self.custom_emails)
                    if available > 0:
                        only_custom = True
                        total = available
                        # reflect actual total in the UI so users see that only their emails will be used
                        try:
                            self.num_accounts.delete(0, tk.END)
                            self.num_accounts.insert(0, str(total))
                        except Exception:
                            pass

            except Exception:
                pass

            self.total_to_create = total
            # created_count == number of successful creations
            self.created_count = 0
            # processed_count == number of email attempts processed (success/fail)
            self.processed_count = 0

            def worker(fast_mode: bool, domain_choice: str, fixed_pwd: str | None, columns_count: int, index: int, only_custom_mode: bool):
                async def run_one():
                    bot = None
                    browser = None

                    # layout/window math (same as before)
                    pad = 10
                    columns = max(1, min(3, columns_count))
                    layout_columns = columns if columns > 1 else 3
                    win_w = int((self.screen_w - (layout_columns + 1) * pad) / layout_columns)
                    target_height = int(self.screen_h * 0.55)
                    win_h = min(target_height, self.screen_h - 2 * pad, 620)
                    win_w = max(win_w, 400)
                    win_h = max(win_h, 355)
                    total_used_w = win_w * layout_columns + pad * (layout_columns + 1)
                    if total_used_w > self.screen_w:
                        win_w = int((self.screen_w - (layout_columns + 1) * pad) / layout_columns)
                    win_h = min(win_h, self.screen_h - 2 * pad)
                    col = index % layout_columns
                    x = pad + col * (win_w + pad)
                    y = pad

                    # pick worker proxy
                    proxy_for_worker = None
                    try:
                        with self.create_proxy_lock:
                            if self.create_proxies:
                                if self.create_proxy_index >= len(self.create_proxies):
                                    self.create_proxy_index = 0
                                proxy_for_worker = self.create_proxies[self.create_proxy_index]
                                self.create_proxy_index = (self.create_proxy_index + 1) % max(1, len(self.create_proxies))
                    except Exception:
                        proxy_for_worker = None

                    try:
                        # init bot/browser
                        bot = await botright.Botright(headless=False, block_images=fast_mode, user_action_layer=False)
                        browser = await bot.new_browser(viewport={"width": win_w, "height": win_h}, extra_args=[f"--window-position={x},{y}", f"--window-size={win_w},{win_h}"], proxy=proxy_for_worker or None)
                        page = await browser.new_page()
                        try:
                            await page.bring_to_front()
                        except Exception:
                            pass
                        await asyncio.sleep(1.0)

                        while self.running:
                            # decide if we stop (different behavior for only_custom_mode)
                            with lock:
                                if only_custom_mode:
                                    if self.processed_count >= self.total_to_create:
                                        break
                                else:
                                    if self.created_count >= self.total_to_create:
                                        break

                            # default per-job assignments
                            assigned_proxy = proxy_for_worker
                            assigned_password = fixed_pwd
                            email_prefix_to_use = "myuser"
                            domain = domain_choice if domain_choice in ("hotmail", "outlook") else random.choice(["hotmail", "outlook"])
                            processed_this_attempt = False

                            # consume a custom entry if available
                            if self.use_custom_emails.get():
                                next_entry = None
                                with self.custom_emails_lock:
                                    if self.custom_emails:
                                        try:
                                            next_entry = self.custom_emails.popleft()
                                        except Exception:
                                            next_entry = None
                                # stop if only_custom_mode is set and no entries left
                                if only_custom_mode and not next_entry:
                                    break
                                if next_entry:
                                    processed_this_attempt = True
                                    if isinstance(next_entry, dict):
                                        ep = (next_entry.get('email') or next_entry.get('Email') or '').strip()
                                        email_part = ep
                                        p = (next_entry.get('proxy') or next_entry.get('proxy_create') or next_entry.get('warmup_proxy') or '').strip() or None
                                        pw = (next_entry.get('password') or next_entry.get('Password') or '').strip() or None
                                    else:
                                        raw = str(next_entry)
                                        parts = [p.strip() for p in raw.split('|',1)] if '|' in raw else ([p.strip() for p in raw.split(',',1)] if ',' in raw else [raw])
                                        email_part = parts[0]
                                        p = parts[1] if len(parts) > 1 else None
                                        pw = None

                                    if '@' in (email_part or ''):
                                        local, dom = email_part.split('@',1)
                                        email_prefix_to_use = local
                                        dom_lower = dom.lower()
                                        if 'hotmail' in dom_lower:
                                            domain = 'hotmail'
                                        elif 'outlook' in dom_lower:
                                            domain = 'outlook'
                                    else:
                                        email_prefix_to_use = email_part or email_prefix_to_use

                                    if p:
                                        assigned_proxy = p
                                    if pw:
                                        assigned_password = pw

                            # if proxy changed for this job, recreate browser/context
                            if assigned_proxy and assigned_proxy != proxy_for_worker:
                                try:
                                    await browser.close()
                                except Exception:
                                    pass
                                try:
                                    browser = await bot.new_browser(viewport={"width": win_w, "height": win_h}, extra_args=[f"--window-position={x},{y}", f"--window-size={win_w},{win_h}"], proxy=assigned_proxy)
                                    page = await browser.new_page()
                                    try:
                                        await page.bring_to_front()
                                    except Exception:
                                        pass
                                except Exception:
                                    print("⚠️ Không thể khởi tạo browser với proxy riêng cho email, tiếp tục với proxy worker (nếu có)")
                                else:
                                    proxy_for_worker = assigned_proxy

                            creator = HotmailAccountCreator()
                            try:
                                account = await creator.create_account(page, email_prefix=email_prefix_to_use, domain=domain, password=assigned_password)
                                success = bool(account)
                                attempt_email = creator.last_full_email or "(unknown)"
                                attempt_password = creator.last_password or "-"
                            except TargetClosedError:
                                success = False
                                attempt_email = creator.last_full_email or "(unknown)"
                                attempt_password = creator.last_password or "-"
                                # If page closed, reduce the target count
                                with lock:
                                    if self.total_to_create > self.created_count:
                                        self.total_to_create -= 1
                            except Exception as e:
                                success = False
                                attempt_email = creator.last_full_email or "(unknown)"
                                attempt_password = creator.last_password or "-"
                                print(f"⚠️ Lỗi trong worker: {e}")

                            # record results and counts
                            with lock:
                                if success:
                                    self.created_count += 1
                                if processed_this_attempt or only_custom_mode:
                                    # count attempts when using custom list
                                    self.processed_count += 1

                            # update UI
                            if success and self.running:
                                self.window.after(0, self.add_account_to_table, account["email"], account["password"], "SUCCESS")
                            else:
                                if self.running:
                                    self.window.after(0, self.add_account_to_table, attempt_email, attempt_password, "FAILED")

                            # persist proxy used for this account if present
                            try:
                                if success and (proxy_for_worker or assigned_proxy):
                                    proxy_to_store = assigned_proxy or proxy_for_worker
                                    try:
                                        update_warmup_status(account["email"], status="created", proxy=proxy_to_store)
                                    except Exception:
                                        pass
                            except Exception:
                                pass

                            # update status text
                            if only_custom_mode:
                                self.window.after(0, self.status_var.set, f"Đã tạo {self.processed_count}/{self.total_to_create} tài khoản")
                                if self.processed_count >= self.total_to_create:
                                    break
                            else:
                                self.window.after(0, self.status_var.set, f"Đã tạo {self.created_count}/{self.total_to_create} tài khoản")
                                if self.created_count >= self.total_to_create:
                                    break

                            # short pause
                            await asyncio.sleep(random.uniform(20,40))

                    finally:
                        try:
                            if browser:
                                await browser.close()
                        except Exception:
                            pass
                        try:
                            if bot:
                                await bot.close()
                        except Exception:
                            pass

                asyncio.run(run_one())

            layout_columns = max(3, concurrency)
            self.threads = []
            for i in range(concurrency):
                if not self.running:
                    break
                t = threading.Thread(target=worker, args=(fast_mode_val, chosen_domain, fixed_pwd_value, layout_columns, i, only_custom), daemon=True)
                self.threads.append(t)
                t.start()
                # Giảm delay để tất cả cửa sổ khởi động gần như cùng lúc
                time.sleep(0.3)

            print(f"Đã bắt đầu tạo {total} tài khoản với tối đa {concurrency} cửa sổ...")
        except ValueError:
            print("Vui lòng nhập số hợp lệ")

    def run(self):
        self.window.mainloop()

    # ====== TÍNH NĂNG NUÔI ======
    def load_warmup_accounts(self):
        if self.warmup_running:
            self.warmup_status_var.set("Đang nuôi, vui lòng dừng trước khi tải lại")
            return
        raw_limit = (self.warmup_limit_entry.get() or "").strip()
        status_filter = (self.warmup_status_filter.get() or "created").lower()

        if status_filter == "all":
            limit = None
        else:
            try:
                limit = int(raw_limit)
                if limit <= 0:
                    raise ValueError
            except (TypeError, ValueError):
                limit = 3
                self.warmup_limit_entry.delete(0, tk.END)
                self.warmup_limit_entry.insert(0, "3")

        display_filter = status_filter if status_filter else "created"
        accounts = fetch_pending_accounts(limit=limit, status_filter=display_filter)
        self.warmup_accounts = accounts
        self.warmup_email_map = {acc["email"]: acc for acc in accounts}
        for item in self.warmup_tree.get_children():
            self.warmup_tree.delete(item)
        for idx, acc in enumerate(accounts, 1):
            last = acc.get("last_activity_at") or "-"
            # try to extract OTPs from warmup_note field if present
            note = acc.get("warmup_note") or ""
            otp_display = "-"
            try:
                # if note contains 'OTPs:' suffix, show it compactly
                if "OTP" in note or "OTPs" in note:
                    otp_display = note
                else:
                    otp_display = "-"
            except Exception:
                otp_display = "-"
            status = acc.get("status", "created")
            self.warmup_tree.insert("", "end", iid=acc["email"], values=(idx, acc["email"], status, last, otp_display))
        if accounts:
            self.warmup_status_var.set(f"Đã nạp {len(accounts)} tài khoản (lọc {display_filter})")
        else:
            self.warmup_status_var.set(f"Không tìm thấy tài khoản status={display_filter}")

    def start_warmup(self):
        if self.warmup_running:
            self.warmup_status_var.set("Đang nuôi rồi")
            return
        mode = self.warmup_mode.get()
        accounts: list[dict] = []
        if mode == "manual":
            selected = self.warmup_tree.selection()
            if not selected:
                self.warmup_status_var.set("Hãy chọn ít nhất 1 tài khoản trong bảng")
                return
            for iid in selected:
                acc = self.warmup_email_map.get(iid)
                if acc:
                    accounts.append(acc)
        else:
            if not self.warmup_accounts:
                self.load_warmup_accounts()
            accounts = list(self.warmup_accounts)

        if not accounts:
            self.warmup_status_var.set("Không có tài khoản để nuôi")
            return

        try:
            requested_conc = int(self.warmup_concurrency_var.get())
        except (TypeError, ValueError):
            requested_conc = 1
        concurrency = max(1, min(3, requested_conc, len(accounts)))

        # If user provided warmup proxy file, assign proxies round-robin to accounts missing proxy
        if self.use_warmup_proxies.get() and self.warmup_proxies:
            try:
                with self.warmup_proxy_lock:
                    idx = self.warmup_proxy_index or 0
                    for acc in accounts:
                        if not acc.get("proxy"):
                            proxy = self.warmup_proxies[idx % len(self.warmup_proxies)]
                            acc["proxy"] = proxy
                            try:
                                update_warmup_status(acc["email"], status="created", proxy=proxy)
                            except Exception:
                                pass
                            idx += 1
                    self.warmup_proxy_index = idx % max(1, len(self.warmup_proxies))
            except Exception:
                pass

        self.warmup_queue = queue.Queue()
        for acc in accounts:
            self.warmup_queue.put(acc)

        self.warmup_total = len(accounts)
        self.warmup_completed = 0
        self.warmup_active_workers = concurrency
        self.warmup_running = True
        self.warmup_worker_threads = []
        self.warmup_email_ring = fetch_all_emails(status_filter="all")
        self.warmup_status_var.set(f"Đang nuôi {self.warmup_total} tài khoản (tối đa {concurrency} cửa sổ)...")

        for i in range(concurrency):
            t = threading.Thread(
                target=self._warmup_worker,
                args=(self.warmup_queue, i, concurrency),
                daemon=True,
            )
            self.warmup_worker_threads.append(t)
            t.start()

    def stop_warmup(self):
        if not self.warmup_running:
            self.warmup_status_var.set("Không có phiên nuôi đang chạy")
            return
        self.warmup_running = False
        if self.warmup_queue:
            try:
                while not self.warmup_queue.empty():
                    self.warmup_queue.get_nowait()
            except queue.Empty:
                pass
        self.warmup_status_var.set("Đang dừng warm-up...")

    def _warmup_worker(self, work_queue: queue.Queue, worker_index: int, columns_count: int):
        async def runner():
            window_conf = self._calc_window_conf(columns_count, worker_index)
            while self.warmup_running:
                try:
                    account = work_queue.get_nowait()
                except queue.Empty:
                    break
                email = account["email"]
                target_email = self._pick_warmup_recipient(email)
                self.window.after(0, self._update_warmup_row, email, "Đang đăng nhập...", "...", "")
                success = False
                note = ""
                try:
                    proxy_to_use = account.get("proxy") if self.use_warmup_proxies.get() else None
                    success, otps = await warmup_account(
                        account,
                        proxy=proxy_to_use,
                        window_conf=window_conf,
                        target_email=target_email,
                        fast=self.warmup_fast.get(),
                    )
                except Exception as exc:
                    note = str(exc)
                    success = False
                status_text = "WARMED" if success else "FAILED"
                last_activity = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S") if success else "-"
                otp_text = ""
                try:
                    if otps:
                        otp_text = ",".join([str(x) for x in otps])
                except Exception:
                    otp_text = ""
                self.window.after(0, self._update_warmup_row, email, status_text, last_activity, otp_text)
                message = f"{email} -> {status_text}"
                if note:
                    message += f" ({note})"
                with self.warmup_lock:
                    self.warmup_completed += 1
                    done = self.warmup_completed
                    total = self.warmup_total
                self.window.after(0, self.warmup_status_var.set, f"{done}/{total}: {message}")
                if not self.warmup_running:
                    break
                if not work_queue.empty():
                    await asyncio.sleep(random.uniform(8, 15))

            with self.warmup_lock:
                self.warmup_active_workers -= 1
                last_worker = self.warmup_active_workers <= 0

            if last_worker:
                self.warmup_running = False
                self.window.after(0, self.warmup_status_var.set, "Đã hoàn tất hoặc dừng nuôi")

        asyncio.run(runner())

    def _pick_warmup_recipient(self, sender_email: str) -> str:
        ring = [email for email in self.warmup_email_ring if email]
        if not ring:
            return sender_email
        if len(ring) == 1:
            return ring[0]
        try:
            idx = ring.index(sender_email)
        except ValueError:
            return ring[0]
        return ring[(idx + 1) % len(ring)]

    def _update_warmup_row(self, email: str, status_text: str, last_activity: str, otp_text: str = ""):
        if email not in self.warmup_tree.get_children():
            return
        current = list(self.warmup_tree.item(email, "values"))
        if not current:
            return
        current[2] = status_text
        current[3] = last_activity
        # Ensure we have an OTP column value available
        try:
            if len(current) < 5:
                # pad if needed
                while len(current) < 5:
                    current.append("")
            current[4] = otp_text or current[4] or "-"
        except Exception:
            pass
        self.warmup_tree.item(email, values=current)

    # ----- File/load helpers -----
    def load_custom_emails(self):
        path = filedialog.askopenfilename(title="Chọn file chứa email (txt/csv/xlsx)", filetypes=[("Excel files", "*.xlsx;*.xls"), ("Text/CSV", "*.txt;*.csv"), ("All files", "*")])
        if not path:
            return
        # if excel but no reader libs installed -> inform user
        if path.lower().endswith(('.xls', '.xlsx')) and not (pd or load_workbook):
            messagebox.showerror("Thiếu thư viện",
                                 "Để đọc file Excel bạn cần cài pandas hoặc openpyxl. Ví dụ: pip install pandas openpyxl")
            return
        try:
            if path.lower().endswith(('.xls', '.xlsx')):
                # Read Excel using pandas
                if pd:
                    df = pd.read_excel(path, dtype=str)
                    df = df.fillna("")
                elif load_workbook:
                    wb = load_workbook(path, read_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        messagebox.showwarning("Không có dữ liệu", "File Excel rỗng")
                        return
                    headers = [str(x) if x is not None else "" for x in rows[0]]
                    data_rows = rows[1:]
                    import collections
                    df = collections.defaultdict(list)
                    for r in data_rows:
                        for i, val in enumerate(r):
                            df[headers[i] if i < len(headers) else str(i)].append(val if val is not None else "")
                    # convert to a simple object with .columns and list-like access used later
                    class SimpleDF:
                        def __init__(self, mapping):
                            self._m = mapping
                            self.columns = list(mapping.keys())
                        def iterrows(self):
                            # return (idx, rowdict)
                            length = len(next(iter(self._m.values())) ) if self._m else 0
                            for idx in range(length):
                                row = {k: (self._m[k][idx] if idx < len(self._m[k]) else "") for k in self._m}
                                yield idx, row
                        def __getitem__(self, col):
                            return self._m.get(col, [])
                    df = SimpleDF(df)
                # Try to find column names for email, password and proxy
                email_col = None
                proxy_col = None
                password_col = None
                warmup_col = None
                cols = [c.lower() for c in df.columns]
                for c in df.columns:
                    low = c.lower()
                    if not email_col and any(k in low for k in ("email", "e-mail", "address")):
                        email_col = c
                    if not password_col and any(k in low for k in ("password", "pwd", "pass")):
                        password_col = c
                    if not proxy_col and any(k in low for k in ("proxy", "proxy_create", "create_proxy")):
                        proxy_col = c
                    if not warmup_col and any(k in low for k in ("warmup_proxy", "proxy_warmup", "proxy_warm")):
                        warmup_col = c

                items = []
                if email_col:
                    for _, row in df.iterrows():
                        email_val = str(row.get(email_col) or "").strip()
                        if not email_val:
                            continue
                        entry = {"email": email_val}
                        if password_col:
                            pw = str(row.get(password_col) or "").strip()
                            if pw:
                                entry["password"] = pw
                        if proxy_col:
                            p = str(row.get(proxy_col) or "").strip()
                            if p:
                                entry["proxy"] = p
                        if warmup_col:
                            wp = str(row.get(warmup_col) or "").strip()
                            if wp:
                                entry["warmup_proxy"] = wp
                        items.append(entry)
                else:
                    messagebox.showwarning("Không có email", "File excel không có cột email phù hợp")
                    return
            else:
                with open(path, "r", encoding="utf-8") as f:
                    lines = [l.strip() for l in f.readlines()]
                items = [l for l in lines if l and not l.startswith("#")]
        except Exception as e:
            messagebox.showerror("Lỗi đọc file", f"Không thể đọc file: {e}")
            return
        if not items:
            messagebox.showwarning("Không có email", "File không chứa email hợp lệ")
            return
        # normalize entries into dicts: if a simple string is present, we'll keep as raw string
        normalized = []
        for entry in items:
            if isinstance(entry, dict):
                normalized.append(entry)
            else:
                raw = str(entry)
                # try to split if contains separators
                if '|' in raw:
                    left, right = raw.split('|', 1)
                    normalized.append({"email": left.strip(), "proxy": right.strip()})
                elif ',' in raw:
                    parts = [p.strip() for p in raw.split(',') if p.strip()]
                    if len(parts) == 1:
                        normalized.append({"email": parts[0]})
                    elif len(parts) == 2:
                        # decide whether second column looks like a proxy or a password
                        second = parts[1]
                        if ':' in second or '@' in second or second.count('.') >= 1:
                            normalized.append({"email": parts[0], "proxy": parts[1]})
                        else:
                            normalized.append({"email": parts[0], "password": parts[1]})
                    else:
                        # more than two columns -> email,password,proxy (best-effort)
                        normalized.append({"email": parts[0], "password": parts[1], "proxy": parts[2]})
                else:
                    normalized.append({"email": raw})

        # enable custom email mode automatically and store
        with self.custom_emails_lock:
            self.custom_emails.clear()
            for item in normalized:
                self.custom_emails.append(item)
        # If user loaded an XLSX but neither pandas nor openpyxl were available, advise them
        if path.lower().endswith(('.xls', '.xlsx')) and not (pd or load_workbook):
            messagebox.showerror("Thiếu thư viện",
                                 "Để đọc file Excel bạn cần cài pandas hoặc openpyxl. Ví dụ: pip install pandas openpyxl")
            return
        # mark checkbox on so user doesn't need to toggle
        try:
            self.use_custom_emails.set(True)
        except Exception:
            pass
        self.custom_emails_label.config(text=f"({len(self.custom_emails)} email)")
        messagebox.showinfo("Đã nạp email", f"Đã nạp {len(self.custom_emails)} email từ file")

    def load_create_proxies(self):
        path = filedialog.askopenfilename(title="Chọn file proxy cho tạo tài khoản", filetypes=[("Proxy list (.txt/.csv/.xlsx)", "*.txt;*.csv;*.xlsx;*.xls"), ("All files", "*")])
        if not path:
            return
        if path.lower().endswith(('.xls', '.xlsx')) and not (pd or load_workbook):
            messagebox.showerror("Thiếu thư viện",
                                 "Để đọc file Excel bạn cần cài pandas hoặc openpyxl. Ví dụ: pip install pandas openpyxl")
            return
        try:
            if path.lower().endswith(('.xls', '.xlsx')):
                if pd:
                    df = pd.read_excel(path, dtype=str)
                    df = df.fillna("")
                elif load_workbook:
                    wb = load_workbook(path, read_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        messagebox.showwarning("Không có dữ liệu", "File Excel rỗng")
                        return
                    headers = [str(x) if x is not None else "" for x in rows[0]]
                    data_rows = rows[1:]
                    import collections
                    df = collections.defaultdict(list)
                    for r in data_rows:
                        for i, val in enumerate(r):
                            df[headers[i] if i < len(headers) else str(i)].append(val if val is not None else "")
                    class SimpleDF:
                        def __init__(self, mapping):
                            self._m = mapping
                            self.columns = list(mapping.keys())
                        def __iter__(self):
                            # iterate rows by index
                            length = len(next(iter(self._m.values())) ) if self._m else 0
                            for idx in range(length):
                                yield {k: (self._m[k][idx] if idx < len(self._m[k]) else "") for k in self._m}
                        def __getitem__(self, col):
                            return self._m.get(col, [])
                    df = SimpleDF(df)
                # pick column that looks like proxy if exists, else first column
                proxy_col = None
                for c in df.columns:
                    if 'proxy' in str(c).lower():
                        proxy_col = c
                        break
                if proxy_col is None:
                    try:
                        if hasattr(df, 'shape') and df.shape[1] >= 1:
                            proxy_col = df.columns[0]
                        elif hasattr(df, 'columns') and len(df.columns) >= 1:
                            proxy_col = df.columns[0]
                    except Exception:
                        # best-effort fallback
                        if hasattr(df, 'columns') and len(df.columns) >= 1:
                            proxy_col = df.columns[0]
                vals = df[proxy_col]
                if hasattr(vals, 'tolist'):
                    vals = vals.tolist()
                proxies = [str(v).strip() for v in vals if str(v).strip()]
            else:
                with open(path, "r", encoding="utf-8") as f:
                    lines = [l.strip() for l in f.readlines()]
                proxies = [l for l in lines if l and not l.startswith("#")]
        except Exception as e:
            messagebox.showerror("Lỗi đọc file", f"Không thể đọc file: {e}")
            return
        with self.create_proxy_lock:
            self.create_proxies = proxies
            self.create_proxy_index = 0
        self.create_proxies_label.config(text=f"({len(self.create_proxies)} proxy)")
        messagebox.showinfo("Đã nạp proxy", f"Đã nạp {len(self.create_proxies)} proxy cho tạo tài khoản")

    def load_warmup_proxies(self):
        path = filedialog.askopenfilename(title="Chọn file proxy cho nuôi", filetypes=[("Proxy list (.txt/.csv/.xlsx)", "*.txt;*.csv;*.xlsx;*.xls"), ("All files", "*")])
        if not path:
            return
        if path.lower().endswith(('.xls', '.xlsx')) and not (pd or load_workbook):
            messagebox.showerror("Thiếu thư viện",
                                 "Để đọc file Excel bạn cần cài pandas hoặc openpyxl. Ví dụ: pip install pandas openpyxl")
            return
        try:
            if path.lower().endswith(('.xls', '.xlsx')):
                if pd:
                    df = pd.read_excel(path, dtype=str)
                    df = df.fillna("")
                elif load_workbook:
                    wb = load_workbook(path, read_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        messagebox.showwarning("Không có dữ liệu", "File Excel rỗng")
                        return
                    headers = [str(x) if x is not None else "" for x in rows[0]]
                    data_rows = rows[1:]
                    import collections
                    df = collections.defaultdict(list)
                    for r in data_rows:
                        for i, val in enumerate(r):
                            df[headers[i] if i < len(headers) else str(i)].append(val if val is not None else "")
                    class SimpleDF:
                        def __init__(self, mapping):
                            self._m = mapping
                            self.columns = list(mapping.keys())
                        def __getitem__(self, col):
                            return self._m.get(col, [])
                    df = SimpleDF(df)
                proxy_col = None
                for c in df.columns:
                    if 'proxy' in str(c).lower():
                        proxy_col = c
                        break
                if proxy_col is None:
                    try:
                        if hasattr(df, 'shape') and df.shape[1] >= 1:
                            proxy_col = df.columns[0]
                        elif hasattr(df, 'columns') and len(df.columns) >= 1:
                            proxy_col = df.columns[0]
                    except Exception:
                        if hasattr(df, 'columns') and len(df.columns) >= 1:
                            proxy_col = df.columns[0]
                vals = df[proxy_col]
                if hasattr(vals, 'tolist'):
                    vals = vals.tolist()
                proxies = [str(v).strip() for v in vals if str(v).strip()]
            else:
                with open(path, "r", encoding="utf-8") as f:
                    lines = [l.strip() for l in f.readlines()]
                proxies = [l for l in lines if l and not l.startswith("#")]
        except Exception as e:
            messagebox.showerror("Lỗi đọc file", f"Không thể đọc file: {e}")
            return
        with self.warmup_proxy_lock:
            self.warmup_proxies = proxies
            self.warmup_proxy_index = 0
        self.warmup_proxies_label.config(text=f"({len(self.warmup_proxies)} proxy)")
        messagebox.showinfo("Đã nạp proxy", f"Đã nạp {len(self.warmup_proxies)} proxy cho nuôi")

def main():
    app = MultiHotmailGUI()
    app.run()

if __name__ == "__main__":
    main()
